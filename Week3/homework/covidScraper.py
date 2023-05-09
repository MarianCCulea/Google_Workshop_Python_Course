from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By

# Setup
driver = webdriver.Chrome()

url_pattern = "https://www.mai.gov.ro/informare-covid-19-grupul-de-comunicare-strategica-{}-martie-ora-13-00-2/"

start_date = 1
end_date = 12
data = {}

for date in range(start_date, end_date):
    url = url_pattern.format(date)
    print(url)

    driver.get(url)

    # Find the first table on the page

    try:
        tables = driver.find_element(by=By.TAG_NAME, value="table")
    except:
        print(f"No table found on {url}. Skipping...")
        continue

    # Extract the data from each table

    rows = tables.find_elements(By.TAG_NAME, "tr")
    for i, row in enumerate(rows):
        if i == 0:
            continue  # skip the first row
        cells = row.find_elements(By.TAG_NAME, "td")
        if len(cells) >= 3:
            city = cells[1].text
            value = cells[2].text
            if city not in data:
                data[city] = [value]
            else:
                data[city].append(value)

# Write the data to an Excel file
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "City"

for i in range(0, end_date):
    ws.cell(row=1, column=i + 2, value=f"{i+1} Martie")


for i, (city, values) in enumerate(data.items()):
    ws.cell(row=i + 2, column=1, value=city)
    for j, value in enumerate(values):
        ws.cell(row=i + 2, column=j + 2, value=value)


wb.save("table_data.xlsx")

# Close the driver
driver.quit()
